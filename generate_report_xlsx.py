#!/usr/bin/env python3
"""Spring B2 Excel 报告生成器 — 每次回测后运行，覆盖更新。无基准对比，纯策略指标。"""
import pandas as pd, numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json, os

ROOT = Path(__file__).parent
BACKTEST = ROOT / 'output' / 'backtest'
OUT = BACKTEST / f'Spring_B2_回测报告_{datetime.now().strftime("%Y%m%d")}.xlsx'

# ── 从回测 daily_values.csv 读取净值, trade_log 读取持仓 ──
def build_daily_from_trade_log():
    tl_path = BACKTEST / 'trade_log.csv'
    if not tl_path.exists():
        raise FileNotFoundError('trade_log.csv 不存在，请先运行回测')

    tl = pd.read_csv(tl_path, parse_dates=['date'])
    tl['symbol'] = tl['symbol'].astype(str).str.zfill(6)

    capital = 200_000  # 策略初始资金

    trade_dates = sorted(tl['date'].unique())
    all_dates = pd.date_range(trade_dates[0], trade_dates[-1], freq='B')

    # 从daily_values提取每日净值(每天最后一条, 过滤异常快照)
    dv_path = BACKTEST / 'daily_values.csv'
    daily_snapshot = {}
    if dv_path.exists():
        dv = pd.read_csv(dv_path, parse_dates=['date'])
        for _, r in dv.iterrows():
            d_key = r['date'].date()
            daily_snapshot[d_key] = {'value': float(r['value']), 'cash': float(r['cash']),
                                     'positions': int(r['positions'])}

    # 检测异常快照: 无15:00 + 无交易 + 日涨跌>15% → 用前日值
    trade_dates_set = set(tl['date'].dt.date)
    fixed_dates = set()
    sorted_dates = sorted(daily_snapshot.keys())
    for i, d in enumerate(sorted_dates):
        if i == 0:
            continue
        prev = daily_snapshot[sorted_dates[i-1]]['value']
        cur = daily_snapshot[d]['value']
        chg = abs(cur / prev - 1) if prev > 0 else 0
        if chg > 0.15 and d not in trade_dates_set:
            daily_snapshot[d] = dict(daily_snapshot[sorted_dates[i-1]])
            fixed_dates.add(d)

    if fixed_dates:
        print(f'  过滤异常快照: {len(fixed_dates)}天 {sorted(fixed_dates)}')

    # 如果有daily_values数据, 直接用
    if daily_snapshot:
        nav_rows = []
        prev_eq = float(capital)
        total_equity = float(capital)
        cash = float(capital)
        n_pos = 0
        daily_ret = 0.0
        for d in all_dates:
            cv = daily_snapshot.get(d.date())
            if cv:
                total_equity = cv['value'] * 2
                daily_ret = (total_equity / prev_eq - 1) if prev_eq > 0 else 0
                n_pos = cv['positions']
                cash = cv['cash'] * 2
                prev_eq = total_equity
            else:
                daily_ret = 0.0
                total_equity = prev_eq  # 非交易日沿用
            nav_rows.append({
                'date': d, 'total_equity': total_equity, 'cash': cash,
                'position_value': total_equity - cash,
                'n_positions': n_pos, 'daily_ret': daily_ret,
                'cum_ret': total_equity / capital - 1,
            })

        # 从trade_log重建每日持仓
        positions = {}
        pos_rows = []
        for d in all_dates:
            if d in set(trade_dates):
                day_trades = tl[tl['date'] == d]
                for _, s in day_trades[day_trades['action'] == 'SELL'].iterrows():
                    sym = str(s['symbol']).zfill(6)
                    if sym in positions: del positions[sym]
                for _, b in day_trades[day_trades['action'] == 'BUY'].iterrows():
                    sym = str(b['symbol']).zfill(6)
                    positions[sym] = {'shares': int(b['shares']), 'side': str(b.get('side', ''))}
            for sym, p in positions.items():
                pos_rows.append({'date': d, 'symbol': sym, 'shares': p['shares'], 'side': p.get('side', '')})

        nav = pd.DataFrame(nav_rows); nav['date'] = pd.to_datetime(nav['date'])
        pos = pd.DataFrame(pos_rows); pos['date'] = pd.to_datetime(pos['date'])
        nav['nav'] = nav['total_equity'] / capital; nav['cum_ret_pct'] = nav['nav'] - 1
        return nav, pos, all_dates

    # fallback: 从trade_log+parquet重建
    # 价格缓存: 按需从parquet读取持仓股日线
    _price_cache = {}
    def _get_price(sym, dt):
        key = (sym, dt.date())
        if key not in _price_cache:
            f = ROOT / 'ML_optimization' / 'features' / f'{sym}.parquet'
            if f.exists():
                df = pd.read_parquet(f, columns=['date', 'close'])
                df['date'] = pd.to_datetime(df['date'])
                for _, r in df.iterrows():
                    _price_cache[(sym, r['date'].date())] = float(r['close'])
        return _price_cache.get(key, None)

    positions = {}  # sym -> {shares, entry_price}
    cash = float(capital)
    nav_rows = []
    pos_rows = []

    prev_equity = float(capital)
    trade_date_set = set(trade_dates)

    for d in all_dates:
        d_date = d.date()
        if d in trade_date_set:
            day_trades = tl[tl['date'] == d]

            # 先卖出
            sells = day_trades[day_trades['action'] == 'SELL']
            for _, s in sells.iterrows():
                sym = str(s['symbol']).zfill(6)
                if sym in positions:
                    px = float(s['price'])
                    cash += int(s['shares']) * px * (1 - 0.0003)
                    del positions[sym]

            # 再买入
            buys = day_trades[day_trades['action'] == 'BUY']
            for _, b in buys.iterrows():
                sym = str(b['symbol']).zfill(6)
                px = float(b['price'])
                cost = float(b['cost']) if pd.notna(b.get('cost')) else int(b['shares']) * px * (1 + 0.0003)
                if cost <= cash:
                    cash -= cost
                    positions[sym] = {'shares': int(b['shares']), 'entry_price': px, 'side': str(b.get('side', ''))}

        # 按当日收盘价计算持仓市值
        pos_val = 0.0
        for sym, p in positions.items():
            px = _get_price(sym, d) or p['entry_price']
            pos_val += p['shares'] * px

        total_equity = cash + pos_val
        daily_ret = (total_equity / prev_equity - 1) if prev_equity > 0 else 0

        nav_rows.append({
            'date': d, 'total_equity': total_equity, 'cash': cash,
            'position_value': pos_val, 'n_positions': len(positions),
            'daily_ret': daily_ret,
            'cum_ret': total_equity / capital - 1,
        })

        for sym, p in positions.items():
            pos_rows.append({
                'date': d, 'symbol': sym, 'shares': p['shares'],
                'side': p.get('side', ''),
            })

        prev_equity = total_equity

    nav = pd.DataFrame(nav_rows)
    nav['date'] = pd.to_datetime(nav['date'])
    pos = pd.DataFrame(pos_rows)
    pos['date'] = pd.to_datetime(pos['date'])

    # 计算累计净值
    nav['nav'] = nav['total_equity'] / capital
    nav['cum_ret_pct'] = nav['nav'] - 1

    return nav, pos, all_dates


nav, pos, dates = build_daily_from_trade_log()

# Styles
hfont = Font(bold=True, color='FFFFFF', size=10)
hfill = PatternFill('solid', fgColor='2F5496')
gfill = PatternFill('solid', fgColor='C6EFCE')
rfill = PatternFill('solid', fgColor='FFC7CE')
bfill = PatternFill('solid', fgColor='BDD7EE')
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
bold = Font(bold=True, size=12)

def hdr(ws, r, cols):
    for c, v in enumerate(cols, 1):
        cl = ws.cell(row=r, column=c, value=v)
        cl.font = hfont; cl.fill = hfill
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        cl.border = border

def auto_w(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 2, 22)

wb = Workbook()

# ═══ Sheet 1: 每日净值 ═══
ws = wb.active; ws.title = '每日净值'
ws.cell(row=1, column=1, value=f'Spring B2 2026年回测 — 每日净值 (至 {dates[-1].date()})').font = bold
cols = ['日期', '总权益', '现金', '持仓市值', '持仓数', '仓位%', '日收益', '累计收益', '净值']
hdr(ws, 3, cols)
for r, (_, row) in enumerate(nav.iterrows(), 4):
    ws.cell(row=r, column=1, value=str(row['date'].date()))
    ws.cell(row=r, column=2, value=int(row['total_equity']))
    ws.cell(row=r, column=3, value=int(row['cash']))
    ws.cell(row=r, column=4, value=int(row['position_value']))
    ws.cell(row=r, column=5, value=int(row['n_positions']))
    pv = row['position_value']; te = row['total_equity']
    if te > 0:
        ws.cell(row=r, column=6, value=round(pv / te * 100, 1)).number_format = '0.0'
    ws.cell(row=r, column=7, value=round(row['daily_ret'] * 100, 2)).number_format = '0.00'
    ws.cell(row=r, column=8, value=round(row['cum_ret'] * 100, 2)).number_format = '0.00'
    ws.cell(row=r, column=9, value=round(row['nav'], 4)).number_format = '0.0000'
    dr = row['daily_ret']
    if dr > 0: ws.cell(row=r, column=7).fill = gfill
    elif dr < 0: ws.cell(row=r, column=7).fill = rfill
    for c in range(1, 10): ws.cell(row=r, column=c).border = border
auto_w(ws)

# ═══ Sheet 2: 每日持仓 ═══
ws = wb.create_sheet('每日持仓')
ws.cell(row=1, column=1, value=f'Spring B2 每日持仓明细 (至 {dates[-1].date()})').font = bold
hdr(ws, 3, ['日期', '代码', '股数', '持仓状态', '持仓类型'])
prev_s = set()
pos_cache = []
for i, d in enumerate(dates):
    dp = pos[pos['date'] == d]; cur = set(dp['symbol'].tolist())
    new_s = cur - prev_s
    nxt = set(pos[pos['date'] == dates[i + 1]]['symbol'].tolist()) if i < len(dates) - 1 else set()
    for _, pr in dp.iterrows():
        t = '★新进' if pr['symbol'] in new_s else ('▼卖出' if pr['symbol'] not in nxt and i < len(dates) - 1 else '持有')
        s = str(pr.get('side', '')).replace('base', '底仓').replace('sat', '卫星').replace('score', '排名')
        pos_cache.append((d, pr['symbol'], pr['shares'], t, s))
    prev_s = cur

for r, (d, sym, sh, t, s) in enumerate(pos_cache, 4):
    ws.cell(row=r, column=1, value=str(d.date())); ws.cell(row=r, column=2, value=sym)
    ws.cell(row=r, column=3, value=sh); ws.cell(row=r, column=4, value=t); ws.cell(row=r, column=5, value=s)
    if t == '★新进': ws.cell(row=r, column=4).fill = gfill
    elif t == '▼卖出': ws.cell(row=r, column=4).fill = rfill
    if '底仓' in s: ws.cell(row=r, column=5).fill = bfill
    for c in range(1, 6): ws.cell(row=r, column=c).border = border
auto_w(ws)

# ═══ Sheet 3: 每日交易摘要 ═══
ws = wb.create_sheet('每日交易摘要')
ws.cell(row=1, column=1, value=f'Spring B2 每日交易摘要 (至 {dates[-1].date()})').font = bold
hdr(ws, 3, ['日期', '持仓数', '新进股', '卖出股', '总权益', '日收益%', '累计收益%'])
prev_s = set()
trade_cache = []
for i, d in enumerate(dates):
    dp = pos[pos['date'] == d]; cur = set(dp['symbol'].tolist())
    new_s = cur - prev_s; exit_s = prev_s - cur
    nr = nav[nav['date'] == d]
    nv = nr.iloc[0].to_dict() if len(nr) > 0 else {}
    trade_cache.append((d, len(cur), new_s, exit_s, nv))
    prev_s = cur

for r, (d, n_cur, new_s, exit_s, nv) in enumerate(trade_cache, 4):
    ws.cell(row=r, column=1, value=str(d.date()))
    ws.cell(row=r, column=2, value=n_cur)
    ws.cell(row=r, column=3, value=','.join(sorted(new_s)) if new_s else '-')
    ws.cell(row=r, column=4, value=','.join(sorted(exit_s)) if exit_s else '-')
    ws.cell(row=r, column=5, value=int(nv.get('total_equity', 0)))
    ws.cell(row=r, column=6, value=round(nv.get('daily_ret', 0) * 100, 2))
    ws.cell(row=r, column=7, value=round(nv.get('cum_ret', 0) * 100, 2))
    dr = nv.get('daily_ret', 0)
    if dr > 0: ws.cell(row=r, column=6).fill = gfill
    elif dr < 0: ws.cell(row=r, column=6).fill = rfill
    for c in range(1, 8): ws.cell(row=r, column=c).border = border
auto_w(ws)

# ═══ Sheet 4: 月度汇总 ═══
ws = wb.create_sheet('月度汇总')
ws.cell(row=1, column=1, value='Spring B2 月度绩效汇总').font = bold
nav['month'] = nav['date'].dt.to_period('M')
monthly = nav.groupby('month').agg(
    start_equity=('total_equity', 'first'),
    end_equity=('total_equity', 'last'),
    avg_positions=('n_positions', 'mean'),
    trading_days=('date', 'count'),
).reset_index()
monthly['return'] = monthly['end_equity'] / monthly['start_equity'] - 1
monthly['month'] = monthly['month'].astype(str)
cols = ['月份', '交易天数', '月初权益', '月末权益', '月收益%', '月均持仓']
hdr(ws, 3, cols)
for r, (_, mr) in enumerate(monthly.iterrows(), 4):
    ws.cell(row=r, column=1, value=mr['month']).border = border
    ws.cell(row=r, column=2, value=int(mr['trading_days'])).border = border
    ws.cell(row=r, column=3, value=int(mr['start_equity'])).border = border
    ws.cell(row=r, column=4, value=int(mr['end_equity'])).border = border
    ws.cell(row=r, column=5, value=round(mr['return'] * 100, 2)).number_format = '0.00'
    ws.cell(row=r, column=5).border = border
    ws.cell(row=r, column=6, value=round(mr['avg_positions'], 1)).border = border
    if mr['return'] > 0: ws.cell(row=r, column=5).fill = gfill
    elif mr['return'] < 0: ws.cell(row=r, column=5).fill = rfill
    for c in range(1, 7): ws.cell(row=r, column=c).border = border
auto_w(ws)

# ═══ Sheet 5: 策略概要 ═══
ws = wb.create_sheet('策略概要')
final_nav = nav['nav'].iloc[-1]
final_cum = nav['cum_ret'].iloc[-1]
peak_nav = nav['nav'].max()
mdd = (nav['nav'] / nav['nav'].cummax() - 1).min()
win_days = (nav['daily_ret'] > 0).sum()
info = [
    ('策略', 'Spring B2'),
    ('回测区间', f'{dates[0].date()} → {dates[-1].date()}'),
    ('交易日', len(dates)),
    ('最终净值', f'{final_nav:.4f}'),
    ('累计收益', f'{final_cum*100:+.2f}%'),
    ('年化收益', f'{(final_nav**(252/len(dates))-1)*100:.2f}%'),
    ('最大回撤', f'{mdd*100:.2f}%'),
    ('胜率(日)', f'{win_days/len(dates)*100:.1f}%'),
    ('月均持仓', f'{nav["n_positions"].mean():.1f} 只'),
    ('总交易', f'{len(pos)//2} 笔'),
    ('更新', datetime.now().strftime('%Y-%m-%d %H:%M')),
]
for r, (k, v) in enumerate(info, 3):
    ws.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws.cell(row=r, column=2, value=v)
ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 40

wb.save(str(OUT))
print(f'已保存: {OUT}')
